"""
Comprehensive Adversarial Stress Test Suite for Frontera Eficiente.

Challenge Areas:
1. Data Loader with Adversarial CSV/Excel/Text inputs
2. Multi-Market Calendar Harmonization & Asynchronous Market Schedules
3. Excel Export Engine (.xlsx generation, OpenPyXL structures, formats, formulas)
4. UI & Presets Pipeline Robustness
"""

from __future__ import annotations

import io
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd
import pytest

from src.data.loader import (
    validate_tickers,
    parse_manual_data,
    load_manual_file,
    _detect_csv_delimiter,
    _detect_and_clean_numeric_col,
)
from src.data.cleaner import (
    normalize_datetime_index,
    align_to_calendar,
    trim_common_inception,
    validate_price_data,
    calculate_daily_returns,
    clean_and_align_prices,
)
from src.export.exporter import (
    export_summary_csv,
    export_weights_csv,
    export_correlation_csv,
    export_wealth_series_csv,
    export_full_excel,
    generate_excel_workbook,
    _convert_metrics_dict_to_df,
    _convert_weights_dict_to_df,
)
from src.presets.portfolio_presets import (
    PRESETS,
    list_presets,
    get_preset,
    get_preset_tickers,
    get_preset_weights,
    get_preset_description,
    _resolve_preset_key,
)
from src.visualization.plots import (
    plot_efficient_frontier,
    plot_asset_allocation,
    plot_allocation_comparison,
    plot_correlation_heatmap,
    plot_covariance_heatmap,
    plot_historical_backtest,
    plot_monte_carlo_cones,
)


# ===========================================================================
# 1. Adversarial Data Loader Tests
# ===========================================================================

class TestAdversarialDataLoader:

    def test_european_semicolon_and_comma_decimal(self):
        """European format with semicolon delimiters and comma decimals."""
        csv_content = (
            "Fecha;ActivoA;ActivoB;ActivoC\n"
            "2023-01-02;100,50;50,25;1.250,75\n"
            "2023-01-03;101,75;51,00;1.260,50\n"
            "2023-01-04;102,00;50,80;1.255,00\n"
            "2023-01-05;103,20;52,10;1.270,30\n"
            "2023-01-06;102,90;51,95;1.268,00\n"
        )
        df = parse_manual_data(csv_content)
        assert isinstance(df.index, pd.DatetimeIndex)
        assert len(df) == 5
        assert set(df.columns) == {"ActivoA", "ActivoB", "ActivoC"}
        assert np.isclose(df.loc["2023-01-02", "ActivoA"], 100.50)
        assert np.isclose(df.loc["2023-01-02", "ActivoC"], 1250.75)
        assert np.isclose(df.loc["2023-01-06", "ActivoB"], 51.95)

    def test_mixed_date_formats_and_whitespace_sorting(self):
        """Dates with mixed formatting (ISO, slash, space) and padded whitespace with chronological sort."""
        csv_content = (
            "  Date  ,  SPY  ,  TLT  \n"
            " 2023-01-02 , 380.50 , 100.20 \n"
            " 2023-01-03 , 382.10 , 99.80 \n"
            " 2023/01/04 , 381.90 , 100.50 \n"
            " 2023-01-05 , 385.00 , 101.10 \n"
            " 2023-01-06 , 389.20 , 100.90 \n"
        )
        df = parse_manual_data(csv_content)
        assert len(df) == 5
        assert "SPY" in df.columns
        assert "TLT" in df.columns
        assert df["SPY"].iloc[0] == 380.50
        assert df["TLT"].iloc[-1] == 100.90
        # Index is sorted chronologically
        assert df.index.is_monotonic_increasing

    def test_pipe_and_tab_delimiters(self):
        """Auto-detection of pipe (|) and tab (\\t) delimiters."""
        pipe_csv = (
            "Date|AAPL|MSFT\n"
            "2023-01-02|150.0|240.0\n"
            "2023-01-03|152.0|242.0\n"
            "2023-01-04|151.0|241.0\n"
            "2023-01-05|153.0|243.0\n"
            "2023-01-06|155.0|245.0\n"
        )
        df_pipe = parse_manual_data(pipe_csv)
        assert set(df_pipe.columns) == {"AAPL", "MSFT"}
        assert len(df_pipe) == 5

        tab_csv = (
            "Date\tAAPL\tMSFT\n"
            "2023-01-02\t150.0\t240.0\n"
            "2023-01-03\t152.0\t242.0\n"
            "2023-01-04\t151.0\t241.0\n"
            "2023-01-05\t153.0\t243.0\n"
            "2023-01-06\t155.0\t245.0\n"
        )
        df_tab = parse_manual_data(tab_csv)
        assert set(df_tab.columns) == {"AAPL", "MSFT"}
        assert len(df_tab) == 5

    def test_garbage_rows_currency_symbols_and_nulls(self):
        """CSV containing currency symbols ($), percentage strings, null markers."""
        csv_content = (
            "Timestamp,Asset1,Asset2\n"
            "2023-01-02,$100.50,50.0%\n"
            "2023-01-03,$102.00,51.5%\n"
            "2023-01-04,#N/A,52.0%\n"
            "2023-01-05,$103.50,NULL\n"
            "2023-01-06,$104.20,53.2%\n"
        )
        df = parse_manual_data(csv_content)
        assert len(df) == 5
        assert df["Asset1"].iloc[0] == 100.50
        assert np.isnan(df["Asset1"].iloc[2])  # #N/A converted to NaN
        assert df["Asset2"].iloc[0] == 50.0
        assert np.isnan(df["Asset2"].iloc[3])  # NULL converted to NaN

    def test_duplicate_date_rows_deduplication(self):
        """Ensure duplicate date rows keep first observation without crash."""
        csv_content = (
            "Date,SPY\n"
            "2023-01-02,380.0\n"
            "2023-01-02,381.0\n"
            "2023-01-03,382.0\n"
            "2023-01-04,383.0\n"
            "2023-01-05,384.0\n"
        )
        df = parse_manual_data(csv_content)
        assert len(df) == 4
        assert df.loc["2023-01-02", "SPY"] == 380.0

    def test_integer_yyyymmdd_dates(self):
        """Integer date representations (e.g. 20230102)."""
        csv_content = (
            "Date,AAPL,MSFT\n"
            "20230102,150.0,240.0\n"
            "20230103,152.5,242.0\n"
            "20230104,151.0,241.5\n"
            "20230105,153.0,243.0\n"
            "20230106,155.0,245.0\n"
        )
        df = parse_manual_data(csv_content)
        assert df.index[0] == pd.Timestamp("2023-01-02")
        assert df.index[-1] == pd.Timestamp("2023-01-06")
        assert len(df) == 5

    def test_long_tidy_format(self):
        """Tidy / Long format table with columns Date, Ticker, Price."""
        csv_content = (
            "Date,Ticker,Price\n"
            "2023-01-02,AAPL,150.0\n"
            "2023-01-02,MSFT,250.0\n"
            "2023-01-03,AAPL,152.0\n"
            "2023-01-03,MSFT,251.0\n"
            "2023-01-04,AAPL,151.5\n"
            "2023-01-04,MSFT,253.0\n"
        )
        df = parse_manual_data(csv_content)
        assert set(df.columns) == {"AAPL", "MSFT"}
        assert len(df) == 3
        assert df.loc["2023-01-02", "AAPL"] == 150.0
        assert df.loc["2023-01-04", "MSFT"] == 253.0

    def test_empty_and_invalid_inputs_raise_errors(self):
        """Ensure empty strings, blank files, and non-date CSVs raise informative ValueErrors."""
        with pytest.raises(ValueError, match="empty"):
            parse_manual_data("")

        with pytest.raises(ValueError, match="empty"):
            parse_manual_data("   \n   \n")

        with pytest.raises(ValueError):
            parse_manual_data("ColA,ColB\n1,2\n3,4\n")

    def test_excel_binary_bytes_parsing(self):
        """Create an in-memory Excel workbook (.xlsx) and verify parse_manual_data correctly parses it."""
        df_orig = pd.DataFrame(
            {
                "Fecha": pd.date_range("2023-01-01", periods=5, freq="D"),
                "BTC": [16500.0, 16600.0, 16800.0, 16750.0, 16900.0],
                "ETH": [1200.0, 1210.0, 1225.0, 1220.0, 1240.0],
            }
        )
        excel_buf = io.BytesIO()
        df_orig.to_excel(excel_buf, index=False)
        excel_bytes = excel_buf.getvalue()

        df_parsed = parse_manual_data(excel_bytes)
        assert len(df_parsed) == 5
        assert set(df_parsed.columns) == {"BTC", "ETH"}
        assert df_parsed.loc["2023-01-01", "BTC"] == 16500.0


# ===========================================================================
# 2. Multi-Market Calendar Harmonization & Asynchronous Schedules
# ===========================================================================

class TestMultiMarketCalendarHarmonization:

    def test_asynchronous_crypto_tradfi_byma_alignment(self):
        """
        Adversarial calendar scenario:
        - Crypto (BTC): 7 days a week (continuous)
        - US Equity (SPY): 5 business days, closed on US MLK day (2023-01-16)
        - Argentine CEDEAR (AAPL.BA): Closed on Argentine Carnival (2023-02-20 & 2023-02-21)
        """
        dates_crypto = pd.date_range("2023-01-01", "2023-02-28", freq="D")
        btc_prices = pd.Series(
            16000.0 + np.cumsum(np.random.normal(50, 100, len(dates_crypto))),
            index=dates_crypto,
            name="BTC-USD",
        )

        dates_us = pd.bdate_range("2023-01-01", "2023-02-28")
        # Drop MLK day (Jan 16, 2023)
        dates_us = dates_us.drop(pd.Timestamp("2023-01-16"))
        spy_prices = pd.Series(
            380.0 + np.cumsum(np.random.normal(0.5, 2.0, len(dates_us))),
            index=dates_us,
            name="SPY",
        )

        dates_ba = pd.bdate_range("2023-01-01", "2023-02-28")
        # Drop Argentine Carnival (Feb 20 & 21, 2023)
        dates_ba = dates_ba.drop([pd.Timestamp("2023-02-20"), pd.Timestamp("2023-02-21")])
        ba_prices = pd.Series(
            5000.0 + np.cumsum(np.random.normal(20, 50, len(dates_ba))),
            index=dates_ba,
            name="AAPL.BA",
        )

        raw_df = pd.concat([btc_prices, spy_prices, ba_prices], axis=1)

        # Run clean_and_align_prices with Business Days freq='B'
        clean_p, returns_p = clean_and_align_prices(raw_df, freq="B")

        # 1. Check index is contiguous business days
        expected_b_idx = pd.bdate_range("2023-01-02", "2023-02-28", name="Date")
        assert (clean_p.index == expected_b_idx).all()

        # 2. Check no NaNs exist
        assert not clean_p.isna().any().any()
        assert not returns_p.isna().any().any()

        # 3. Check MLK Day (Jan 16, 2023): SPY should forward-fill Jan 13 price
        jan_13_price = clean_p.loc["2023-01-13", "SPY"]
        jan_16_price = clean_p.loc["2023-01-16", "SPY"]
        assert jan_16_price == jan_13_price
        # On holiday with forward-fill, return should be exactly 0.0
        assert np.isclose(returns_p.loc["2023-01-16", "SPY"], 0.0)

        # 4. Check Carnival Days (Feb 20 & 21, 2023): AAPL.BA should forward-fill Feb 17 price
        feb_17_price = clean_p.loc["2023-02-17", "AAPL.BA"]
        feb_20_price = clean_p.loc["2023-02-20", "AAPL.BA"]
        feb_21_price = clean_p.loc["2023-02-21", "AAPL.BA"]
        assert feb_20_price == feb_17_price
        assert feb_21_price == feb_17_price
        assert np.isclose(returns_p.loc["2023-02-20", "AAPL.BA"], 0.0)
        assert np.isclose(returns_p.loc["2023-02-21", "AAPL.BA"], 0.0)

    def test_no_future_leakage_in_forward_fill(self):
        """Empirically test that forward-fill never leaks future prices into past dates."""
        dates = pd.date_range("2023-01-01", periods=10, freq="B")
        df = pd.DataFrame(
            {
                "AssetA": [10.0, np.nan, np.nan, 20.0, np.nan, 30.0, np.nan, np.nan, 40.0, 50.0],
                "AssetB": [100.0, 101.0, 102.0, 103.0, 104.0, 105.0, 106.0, 107.0, 108.0, 109.0],
            },
            index=dates,
        )

        aligned = align_to_calendar(df, freq="B", method="ffill")
        # AssetA index 1 and 2 must be 10.0, NEVER 20.0
        assert aligned["AssetA"].iloc[1] == 10.0
        assert aligned["AssetA"].iloc[2] == 10.0
        assert aligned["AssetA"].iloc[4] == 20.0
        assert aligned["AssetA"].iloc[6] == 30.0
        assert aligned["AssetA"].iloc[7] == 30.0

    def test_staggered_asset_inception_trimming(self):
        """Asset with inception in 2022 vs Asset with inception in 2020."""
        dates = pd.date_range("2020-01-01", "2023-01-01", freq="B")
        df = pd.DataFrame(index=dates)
        df["SPY"] = 300.0 + np.linspace(0, 100, len(dates))
        # New asset launched on 2022-01-03
        df["NEW_CO"] = np.nan
        df.loc["2022-01-03":, "NEW_CO"] = 50.0 + np.linspace(0, 10, len(df.loc["2022-01-03":]))

        trimmed = trim_common_inception(df)
        assert trimmed.index.min() == pd.Timestamp("2022-01-03")
        assert not trimmed.isna().any().any()

    def test_completely_non_overlapping_assets_raises(self):
        """Two assets with zero overlap in their observation periods must raise ValueError."""
        dates_a = pd.date_range("2015-01-01", "2018-01-01", freq="B")
        dates_b = pd.date_range("2019-01-01", "2022-01-01", freq="B")
        df_a = pd.Series(100.0, index=dates_a, name="A")
        df_b = pd.Series(200.0, index=dates_b, name="B")
        df = pd.concat([df_a, df_b], axis=1)

        with pytest.raises(ValueError, match="No common overlapping date range"):
            trim_common_inception(df)

    def test_flat_series_zero_variance_raises(self):
        """Stale or flat price series must be caught by validate_price_data."""
        dates = pd.date_range("2023-01-01", periods=10, freq="B")
        df = pd.DataFrame(
            {
                "Active": [10.0, 10.2, 10.1, 10.3, 10.4, 10.5, 10.6, 10.7, 10.8, 10.9],
                "Flat": [100.0] * 10,
            },
            index=dates,
        )
        with pytest.raises(ValueError, match="near-zero return variance"):
            validate_price_data(df)


# ===========================================================================
# 3. Excel Export Engine & OpenPyXL Structure Stress-Testing
# ===========================================================================

class TestExcelExportEngine:

    def test_full_excel_workbook_sheets_and_formatting(self):
        """
        Verify that export_full_excel produces a valid multi-sheet workbook with:
        - Correct sheet names
        - Navy headers with white bold font
        - Correct number formats for percentages, ratios, and currencies
        - Dynamic =SUM() formulas in Ponderaciones total check row
        """
        metrics_data = {
            "Usuario": {
                "annualized_return": 0.145,
                "cagr": 0.138,
                "annualized_volatility": 0.162,
                "sharpe_ratio": 0.648,
                "sortino_ratio": 0.912,
                "calmar_ratio": 0.854,
                "max_drawdown": -0.161,
                "var_95_hist": -0.0152,
                "cvar_95_hist": -0.0234,
                "recovery_days": 142,
            },
            "Max Sharpe": {
                "annualized_return": 0.182,
                "cagr": 0.171,
                "annualized_volatility": 0.155,
                "sharpe_ratio": 0.916,
                "sortino_ratio": 1.340,
                "calmar_ratio": 1.425,
                "max_drawdown": -0.120,
                "var_95_hist": -0.0131,
                "cvar_95_hist": -0.0195,
                "recovery_days": 85,
            },
            "GMV": {
                "annualized_return": 0.082,
                "cagr": 0.079,
                "annualized_volatility": 0.098,
                "sharpe_ratio": 0.428,
                "sortino_ratio": 0.612,
                "calmar_ratio": 0.987,
                "max_drawdown": -0.080,
                "var_95_hist": -0.0085,
                "cvar_95_hist": -0.0120,
                "recovery_days": 45,
            },
        }

        weights_data = {
            "SPY": {"Usuario": 0.40, "Max Sharpe": 0.60, "GMV": 0.15},
            "TLT": {"Usuario": 0.30, "Max Sharpe": 0.10, "GMV": 0.65},
            "GLD": {"Usuario": 0.30, "Max Sharpe": 0.30, "GMV": 0.20},
        }

        tickers = ["SPY", "TLT", "GLD"]
        corr_df = pd.DataFrame(
            [[1.0, -0.35, 0.10], [-0.35, 1.0, 0.25], [0.10, 0.25, 1.0]],
            index=tickers,
            columns=tickers,
        )
        cov_df = pd.DataFrame(
            [[0.040, -0.012, 0.003], [-0.012, 0.025, 0.005], [0.003, 0.005, 0.020]],
            index=tickers,
            columns=tickers,
        )
        wealth_df = pd.DataFrame(
            {
                "Usuario": [10000.0, 10150.0, 10220.0, 10180.0, 10350.0],
                "Max Sharpe": [10000.0, 10200.0, 10310.0, 10290.0, 10480.0],
            },
            index=pd.date_range("2023-01-01", periods=5, freq="B"),
        )
        mc_df = pd.DataFrame(
            {
                "Sim_1": [10000.0, 10200.0, 10400.0],
                "Sim_2": [10000.0, 9900.0, 10100.0],
            }
        )

        excel_bytes = export_full_excel(
            metrics_dict=metrics_data,
            weights_dict=weights_data,
            corr_matrix=corr_df,
            cov_matrix=cov_df,
            wealth_df=wealth_df,
            mc_samples_df=mc_df,
        )

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Load back into openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
        sheet_names = wb.sheetnames
        expected_sheets = [
            "Resumen de Métricas",
            "Ponderaciones",
            "Matriz de Correlación",
            "Matriz de Covarianza",
            "Evolución Histórica",
            "Simulación Monte Carlo",
        ]
        for s in expected_sheets:
            assert s in sheet_names

        # Inspect Sheet 1 (Metrics)
        ws_m = wb["Resumen de Métricas"]
        assert ws_m.cell(row=1, column=1).value == "Métrica"
        assert ws_m.cell(row=2, column=2).number_format == "0.00%"

        # Inspect Sheet 2 (Weights)
        ws_w = wb["Ponderaciones"]
        assert ws_w.cell(row=1, column=1).value == "Ticker"
        # Find TOTAL check row (row 5: header=1, SPY=2, TLT=3, GLD=4, TOTAL=5)
        tot_row = len(tickers) + 2
        assert ws_w.cell(row=tot_row, column=1).value == "TOTAL"
        # Check SUM formula in column B, C, D
        assert ws_w.cell(row=tot_row, column=2).value == "=SUM(B2:B4)"
        assert ws_w.cell(row=tot_row, column=3).value == "=SUM(C2:C4)"
        assert ws_w.cell(row=tot_row, column=4).value == "=SUM(D2:D4)"
        assert ws_w.cell(row=tot_row, column=2).number_format == "0.00%"

        # Inspect Sheet 5 (Wealth)
        ws_h = wb["Evolución Histórica"]
        assert ws_h.cell(row=2, column=2).number_format == "$#,##0.00"

    def test_excel_large_asset_universe_formulas(self):
        """Verify formula generation on large asset universes (e.g. 30 assets)."""
        tickers = [f"Asset_{i+1:02d}" for i in range(30)]
        weights_dict = {t: {"Portfolio_A": 1.0 / 30.0, "Portfolio_B": 1.0 / 30.0} for t in tickers}
        excel_bytes = export_full_excel(weights_dict=weights_dict)
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
        ws = wb["Ponderaciones"]
        tot_row = 32  # 1 header + 30 assets + 1 total
        assert ws.cell(row=tot_row, column=1).value == "TOTAL"
        assert ws.cell(row=tot_row, column=2).value == "=SUM(B2:B31)"
        assert ws.cell(row=tot_row, column=3).value == "=SUM(C2:C31)"

    def test_excel_export_edge_cases_empty_or_single_item(self):
        """Excel generation with empty / None / single-row datasets."""
        excel_bytes = export_full_excel(
            metrics_dict=None,
            weights_dict=None,
            corr_matrix=None,
            cov_matrix=None,
            wealth_df=None,
            mc_samples_df=None,
        )
        assert isinstance(excel_bytes, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        assert "Resumen de Métricas" in wb.sheetnames
        assert "Ponderaciones" in wb.sheetnames

    def test_csv_exporters_integrity(self):
        """Test all CSV exporters produce non-empty, comma-separated strings."""
        # Summary CSV
        metrics = {"Usuario": {"annualized_return": 0.12, "sharpe_ratio": 0.8}}
        csv_sum = export_summary_csv(metrics)
        assert "Métrica" in csv_sum
        assert "Usuario" in csv_sum

        # Weights CSV
        weights = {"AAPL": {"Usuario": 0.60}, "MSFT": {"Usuario": 0.40}}
        csv_w = export_weights_csv(weights)
        assert "Ticker" in csv_w
        assert "AAPL" in csv_w

        # Correlation CSV
        corr = pd.DataFrame([[1.0, 0.2], [0.2, 1.0]], index=["AAPL", "MSFT"], columns=["AAPL", "MSFT"])
        csv_corr = export_correlation_csv(corr)
        assert "AAPL" in csv_corr


# ===========================================================================
# 4. UI Presets & Visualization Robustness
# ===========================================================================

class TestPresetsAndVisualizationRobustness:

    def test_all_canonical_presets_validity(self):
        """Verify all 5 canonical presets load correctly and have valid weights summing to 1.0."""
        expected_presets = ["classic_60_40", "all_weather", "big_tech", "cedears_argentina", "crypto_tradfi"]
        for p_key in expected_presets:
            p = get_preset(p_key)
            assert "name" in p
            assert "tickers" in p
            assert "weights" in p
            tickers = p["tickers"]
            weights = p["weights"]
            assert len(tickers) == len(weights)
            assert np.isclose(sum(weights.values()), 1.0, atol=1e-5)

    def test_preset_aliases_resolution(self):
        """Verify alias resolution for Spanish and informal queries."""
        assert _resolve_preset_key("Clásico 60/40") == "classic_60_40"
        assert _resolve_preset_key("60/40") == "classic_60_40"
        assert _resolve_preset_key("ray dalio") == "all_weather"
        assert _resolve_preset_key("CEDEARs") == "cedears_argentina"
        assert _resolve_preset_key("cripto + tradfi") == "crypto_tradfi"

    def test_plotly_figures_edge_cases(self):
        """Verify Plotly figure builders handle minimal inputs without crashing."""
        # 1. Frontier Plot with None inputs
        fig1 = plot_efficient_frontier(rf=0.04)
        assert fig1 is not None

        # 2. Donut Allocation with single asset
        fig2 = plot_asset_allocation({"SPY": 1.0})
        assert fig2 is not None

        # 3. Correlation heatmap 2x2
        corr = pd.DataFrame([[1.0, 0.5], [0.5, 1.0]], index=["A", "B"], columns=["A", "B"])
        fig3 = plot_correlation_heatmap(corr)
        assert fig3 is not None

        # 4. Wealth series backtest
        dates = pd.date_range("2023-01-01", periods=10, freq="B")
        wealth_dict = {
            "Usuario": pd.Series(10000.0 * np.cumprod(1 + np.random.normal(0, 0.01, 10)), index=dates)
        }
        fig4 = plot_historical_backtest(wealth_dict)
        assert fig4 is not None
