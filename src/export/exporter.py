"""
Export Engine Module for Frontera Eficiente.

Provides serialization and export capabilities:
1. CSV Exporters:
   - Summary Metrics (`export_summary_csv`, `export_metrics_csv`)
   - Asset Allocations (`export_weights_csv`)
   - Correlation Matrix (`export_correlation_csv`)
   - Wealth Time Series (`export_wealth_series_csv`)
2. Multi-Sheet Styled Excel Workbook (`export_full_excel`, `generate_excel_workbook`):
   - Sheet 1: Resumen de Métricas
   - Sheet 2: Ponderaciones
   - Sheet 3: Matriz de Correlación
   - Sheet 4: Matriz de Covarianza
   - Sheet 5: Evolución Histórica
   - Sheet 6: Simulación Monte Carlo
   - Navy Header (`#1F4E79`), White bold text, Zebra striping, explicit number formats, auto column widths.
"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Union

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ===========================================================================
# 1. CSV Serialization Helpers
# ===========================================================================

def export_summary_csv(df_or_dict: Union[pd.DataFrame, Dict[str, Any]]) -> str:
    """
    Export summary metrics table to CSV formatted string.

    Parameters
    ----------
    df_or_dict : pd.DataFrame | dict
        Metrics data structure.

    Returns
    -------
    str
        CSV string representation.
    """
    if isinstance(df_or_dict, pd.DataFrame):
        df = df_or_dict
    elif isinstance(df_or_dict, dict):
        df = _convert_metrics_dict_to_df(df_or_dict)
    else:
        raise TypeError(f"Expected pd.DataFrame or dict, got {type(df_or_dict)}")

    return df.to_csv(index=False)


def export_metrics_csv(df_or_dict: Union[pd.DataFrame, Dict[str, Any]]) -> str:
    """Alias for export_summary_csv."""
    return export_summary_csv(df_or_dict)


def export_weights_csv(df_or_dict: Union[pd.DataFrame, Dict[str, Any]]) -> str:
    """
    Export asset weights allocation table to CSV formatted string.

    Parameters
    ----------
    df_or_dict : pd.DataFrame | dict
        Weights data structure.

    Returns
    -------
    str
        CSV string representation.
    """
    if isinstance(df_or_dict, pd.DataFrame):
        df = df_or_dict
    elif isinstance(df_or_dict, dict):
        df = _convert_weights_dict_to_df(df_or_dict)
    else:
        raise TypeError(f"Expected pd.DataFrame or dict, got {type(df_or_dict)}")

    return df.to_csv(index=False)


def export_correlation_csv(corr_df: Union[pd.DataFrame, np.ndarray], tickers: Optional[List[str]] = None) -> str:
    """
    Export correlation matrix to CSV formatted string.

    Parameters
    ----------
    corr_df : pd.DataFrame | np.ndarray
        Correlation matrix.
    tickers : Optional[List[str]], optional
        Ticker names if corr_df is np.ndarray.

    Returns
    -------
    str
        CSV string.
    """
    if isinstance(corr_df, np.ndarray):
        cols = tickers if tickers else [f"Asset_{i+1}" for i in range(corr_df.shape[0])]
        df = pd.DataFrame(corr_df, index=cols, columns=cols)
    else:
        df = corr_df

    return df.to_csv(index=True)


def export_wealth_series_csv(wealth_df: Union[pd.DataFrame, pd.Series]) -> str:
    """
    Export historical wealth series to CSV formatted string.

    Parameters
    ----------
    wealth_df : pd.DataFrame | pd.Series
        Historical wealth series.

    Returns
    -------
    str
        CSV string.
    """
    if isinstance(wealth_df, pd.Series):
        df = wealth_df.to_frame()
    else:
        df = wealth_df

    return df.to_csv(index=True)


# ===========================================================================
# 2. DataFrame Transformation Helpers
# ===========================================================================

_METRIC_LABELS_ES = {
    "annualized_return": "Retorno Anualizado (Aritmético)",
    "cagr": "Retorno Compuesto Anual (CAGR)",
    "annualized_volatility": "Volatilidad Anualizada (Riesgo)",
    "sharpe_ratio": "Ratio de Sharpe",
    "sortino_ratio": "Ratio de Sortino",
    "calmar_ratio": "Ratio de Calmar",
    "max_drawdown": "Máximo Drawdown (MDD)",
    "var_95_hist": "VaR 95% Histórico (1 Día)",
    "var_95_param": "VaR 95% Paramétrico (1 Día)",
    "cvar_95_hist": "CVaR 95% Histórico (Expected Shortfall)",
    "cvar_95_param": "CVaR 95% Paramétrico",
    "recovery_days": "Días de Recuperación de Drawdown",
}


def _convert_metrics_dict_to_df(metrics_input: Union[Dict[str, Any], pd.DataFrame]) -> pd.DataFrame:
    """Convert arbitrary nested metrics dictionary or DataFrame into standard summary DataFrame."""
    if isinstance(metrics_input, pd.DataFrame):
        return metrics_input.copy()

    if not metrics_input:
        return pd.DataFrame(columns=["Métrica"])

    first_key = next(iter(metrics_input.keys()))
    first_val = metrics_input[first_key]

    # Structure 1: Metric -> {Portfolio: Value} e.g. {"Retorno": {"Usuario": 0.15, "Max Sharpe": 0.19}}
    if isinstance(first_val, dict) and not any(k in _METRIC_LABELS_ES for k in first_val.keys()):
        rows = []
        for metric_name, port_map in metrics_input.items():
            row = {"Métrica": metric_name}
            row.update(port_map)
            rows.append(row)
        return pd.DataFrame(rows)

    # Structure 2: Portfolio -> MetricsObj or {metric_name: val} e.g. {"Max Sharpe": {...}, "GMV": {...}}
    portfolios = list(metrics_input.keys())
    # Collect all metric keys
    all_metric_keys = []
    for port in portfolios:
        val_obj = metrics_input[port]
        keys = val_obj.keys() if hasattr(val_obj, "keys") else (val_obj.to_dict().keys() if hasattr(val_obj, "to_dict") else [])
        for k in keys:
            if k not in ("daily_returns", "cumulative_wealth", "drawdown_series") and k not in all_metric_keys:
                all_metric_keys.append(k)

    if not all_metric_keys:
        all_metric_keys = list(_METRIC_LABELS_ES.keys())

    rows = []
    for m_key in all_metric_keys:
        label = _METRIC_LABELS_ES.get(m_key, str(m_key))
        row = {"Métrica": label}
        for port in portfolios:
            val_obj = metrics_input[port]
            if hasattr(val_obj, "__getitem__") and m_key in val_obj:
                row[port] = val_obj[m_key]
            elif hasattr(val_obj, m_key):
                row[port] = getattr(val_obj, m_key)
            else:
                row[port] = np.nan
        rows.append(row)

    return pd.DataFrame(rows)


def _convert_weights_dict_to_df(weights_input: Union[Dict[str, Any], pd.DataFrame]) -> pd.DataFrame:
    """Convert arbitrary weights dictionary or DataFrame into standard weights DataFrame."""
    if isinstance(weights_input, pd.DataFrame):
        if "Ticker" not in weights_input.columns and weights_input.index.name != "Ticker":
            df = weights_input.reset_index()
            if "index" in df.columns:
                df = df.rename(columns={"index": "Ticker"})
            return df
        return weights_input.copy()

    if not weights_input:
        return pd.DataFrame(columns=["Ticker"])

    first_key = next(iter(weights_input.keys()))
    first_val = weights_input[first_key]

    # Structure 1: Ticker -> {Portfolio: weight} e.g. {"AAPL": {"Usuario": 0.33, "Max Sharpe": 0.50}}
    if isinstance(first_val, dict):
        rows = []
        for ticker, port_map in weights_input.items():
            row = {"Ticker": ticker}
            row.update(port_map)
            rows.append(row)
        return pd.DataFrame(rows)

    # Structure 2: Portfolio -> {Ticker: weight} e.g. {"Usuario": {"AAPL": 0.33, "MSFT": 0.33}}
    if isinstance(first_val, (dict, pd.Series)):
        all_tickers = []
        for port, w_dict in weights_input.items():
            for t in w_dict.keys():
                if t not in all_tickers:
                    all_tickers.append(t)

        rows = []
        for ticker in all_tickers:
            row = {"Ticker": ticker}
            for port, w_dict in weights_input.items():
                row[port] = w_dict.get(ticker, 0.0) if isinstance(w_dict, dict) else (w_dict[ticker] if ticker in w_dict else 0.0)
            rows.append(row)
        return pd.DataFrame(rows)

    # Flat dictionary: Ticker -> weight
    rows = [{"Ticker": k, "Peso": v} for k, v in weights_input.items()]
    return pd.DataFrame(rows)


# ===========================================================================
# 3. Multi-Sheet Styled Excel Workbook Generator
# ===========================================================================

def export_full_excel(
    metrics_dict: Optional[Union[Dict[str, Any], pd.DataFrame]] = None,
    weights_dict: Optional[Union[Dict[str, Any], pd.DataFrame]] = None,
    corr_matrix: Optional[pd.DataFrame] = None,
    cov_matrix: Optional[pd.DataFrame] = None,
    wealth_df: Optional[Union[pd.DataFrame, pd.Series]] = None,
    mc_samples_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """
    Generate a professional multi-sheet formatted Excel workbook (.xlsx).

    Styling Standards:
    - Navy Headers (#1F4E79) with bold white font and centered alignment
    - Alternating light zebra striping (#FFFFFF and #F8F9FA)
    - Thin borders (#D9D9D9) on all data cells
    - Explicit number formatting:
      * Percentages (0.00%) for returns, vols, drawdowns, weights, VaR, CVaR
      * Ratios (0.000) for Sharpe, Sortino, Calmar
      * Currency ($#,##0.00) for wealth indices
      * 4-6 decimals for correlation and covariance matrices
    - Automatic column widths with padding

    Sheets:
    1. Resumen de Métricas
    2. Ponderaciones
    3. Matriz de Correlación
    4. Matriz de Covarianza
    5. Evolución Histórica
    6. Simulación Monte Carlo

    Returns
    -------
    bytes
        In-memory Excel binary workbook.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    zebra_fill_gray = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    total_fill = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    bold_font = Font(name="Calibri", size=10, bold=True, color="000000")
    regular_font = Font(name="Calibri", size=10, color="000000")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    thick_bottom_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="double", color="1F4E79"),
    )

    # -----------------------------------------------------------------------
    # Sheet 1: Resumen de Métricas
    # -----------------------------------------------------------------------
    ws_metrics = wb.create_sheet(title="Resumen de Métricas")
    if metrics_dict is not None:
        df_m = _convert_metrics_dict_to_df(metrics_dict)
    else:
        df_m = pd.DataFrame({"Métrica": ["Sin datos"]})

    _write_styled_dataframe(
        ws=ws_metrics,
        df=df_m,
        navy_fill=navy_fill,
        header_font=header_font,
        zebra_fill_white=zebra_fill_white,
        zebra_fill_gray=zebra_fill_gray,
        thin_border=thin_border,
        regular_font=regular_font,
        is_metrics_sheet=True,
    )

    # -----------------------------------------------------------------------
    # Sheet 2: Ponderaciones
    # -----------------------------------------------------------------------
    ws_weights = wb.create_sheet(title="Ponderaciones")
    if weights_dict is not None:
        df_w = _convert_weights_dict_to_df(weights_dict)
    else:
        df_w = pd.DataFrame({"Ticker": ["Sin datos"], "Peso": [0.0]})

    _write_styled_dataframe(
        ws=ws_weights,
        df=df_w,
        navy_fill=navy_fill,
        header_font=header_font,
        zebra_fill_white=zebra_fill_white,
        zebra_fill_gray=zebra_fill_gray,
        thin_border=thin_border,
        regular_font=regular_font,
        is_weights_sheet=True,
        total_fill=total_fill,
        bold_font=bold_font,
        thick_bottom_border=thick_bottom_border,
    )

    # -----------------------------------------------------------------------
    # Sheet 3: Matriz de Correlación
    # -----------------------------------------------------------------------
    ws_corr = wb.create_sheet(title="Matriz de Correlación")
    if corr_matrix is not None:
        df_corr = corr_matrix.reset_index().rename(columns={"index": "Activo"}) if isinstance(corr_matrix, pd.DataFrame) else pd.DataFrame(corr_matrix)
    else:
        df_corr = pd.DataFrame({"Activo": ["N/A"]})

    _write_styled_dataframe(
        ws=ws_corr,
        df=df_corr,
        navy_fill=navy_fill,
        header_font=header_font,
        zebra_fill_white=zebra_fill_white,
        zebra_fill_gray=zebra_fill_gray,
        thin_border=thin_border,
        regular_font=regular_font,
        number_format="0.0000",
    )

    # -----------------------------------------------------------------------
    # Sheet 4: Matriz de Covarianza
    # -----------------------------------------------------------------------
    ws_cov = wb.create_sheet(title="Matriz de Covarianza")
    if cov_matrix is not None:
        df_cov = cov_matrix.reset_index().rename(columns={"index": "Activo"}) if isinstance(cov_matrix, pd.DataFrame) else pd.DataFrame(cov_matrix)
    else:
        df_cov = pd.DataFrame({"Activo": ["N/A"]})

    _write_styled_dataframe(
        ws=ws_cov,
        df=df_cov,
        navy_fill=navy_fill,
        header_font=header_font,
        zebra_fill_white=zebra_fill_white,
        zebra_fill_gray=zebra_fill_gray,
        thin_border=thin_border,
        regular_font=regular_font,
        number_format="0.000000",
    )

    # -----------------------------------------------------------------------
    # Sheet 5: Evolución Histórica (Optional)
    # -----------------------------------------------------------------------
    if wealth_df is not None:
        ws_hist = wb.create_sheet(title="Evolución Histórica")
        if isinstance(wealth_df, pd.Series):
            df_h = wealth_df.to_frame(name="Cartera").reset_index()
        elif isinstance(wealth_df, pd.DataFrame):
            df_h = wealth_df.reset_index()
        else:
            df_h = pd.DataFrame({"Fecha": [], "Valor": []})

        if "index" in df_h.columns:
            df_h = df_h.rename(columns={"index": "Fecha"})
        if "Date" in df_h.columns:
            df_h = df_h.rename(columns={"Date": "Fecha"})

        _write_styled_dataframe(
            ws=ws_hist,
            df=df_h,
            navy_fill=navy_fill,
            header_font=header_font,
            zebra_fill_white=zebra_fill_white,
            zebra_fill_gray=zebra_fill_gray,
            thin_border=thin_border,
            regular_font=regular_font,
            is_wealth_sheet=True,
        )

    # -----------------------------------------------------------------------
    # Sheet 6: Simulación Monte Carlo (Optional)
    # -----------------------------------------------------------------------
    if mc_samples_df is not None:
        ws_mc = wb.create_sheet(title="Simulación Monte Carlo")
        _write_styled_dataframe(
            ws=ws_mc,
            df=mc_samples_df.head(1000),
            navy_fill=navy_fill,
            header_font=header_font,
            zebra_fill_white=zebra_fill_white,
            zebra_fill_gray=zebra_fill_gray,
            thin_border=thin_border,
            regular_font=regular_font,
        )

    # Serialize to memory buffer
    output_buf = io.BytesIO()
    wb.save(output_buf)
    return output_buf.getvalue()


def generate_excel_workbook(
    metrics_df: Optional[Union[Dict[str, Any], pd.DataFrame]] = None,
    weights_df: Optional[Union[Dict[str, Any], pd.DataFrame]] = None,
    corr_df: Optional[pd.DataFrame] = None,
    cov_df: Optional[pd.DataFrame] = None,
    wealth_df: Optional[Union[pd.DataFrame, pd.Series]] = None,
    mc_samples_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """Alias matching canonical parameter names for multi-sheet Excel generator."""
    return export_full_excel(
        metrics_dict=metrics_df,
        weights_dict=weights_df,
        corr_matrix=corr_df,
        cov_matrix=cov_df,
        wealth_df=wealth_df,
        mc_samples_df=mc_samples_df,
    )


# ===========================================================================
# 4. Internal Table Styling Engine
# ===========================================================================

def _write_styled_dataframe(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    df: pd.DataFrame,
    navy_fill: PatternFill,
    header_font: Font,
    zebra_fill_white: PatternFill,
    zebra_fill_gray: PatternFill,
    thin_border: Border,
    regular_font: Font,
    number_format: Optional[str] = None,
    is_metrics_sheet: bool = False,
    is_weights_sheet: bool = False,
    is_wealth_sheet: bool = False,
    total_fill: Optional[PatternFill] = None,
    bold_font: Optional[Font] = None,
    thick_bottom_border: Optional[Border] = None,
) -> None:
    """Helper to write and style a DataFrame onto an OpenPyXL Worksheet."""
    ws.views.sheetView[0].showGridLines = True

    # 1. Header Row
    headers = list(df.columns)
    ws.row_dimensions[1].height = 26
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 2. Data Rows
    num_rows = len(df)
    for r_idx in range(num_rows):
        row_num = r_idx + 2
        ws.row_dimensions[row_num].height = 20
        fill = zebra_fill_white if r_idx % 2 == 0 else zebra_fill_gray
        metric_name = str(df.iloc[r_idx, 0]).lower() if (is_metrics_sheet and len(headers) > 0) else ""

        for c_idx, col_name in enumerate(headers, start=1):
            val = df.iloc[r_idx, c_idx - 1]
            cell = ws.cell(row=row_num, column=c_idx)
            cell.fill = fill
            cell.font = regular_font
            cell.border = thin_border

            # Check for Datetime / Timestamp objects
            if isinstance(val, (pd.Timestamp, np.datetime64)):
                cell.value = str(pd.to_datetime(val).strftime("%Y-%m-%d"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # String column (e.g. Metric name, Ticker, Date) or missing
            if c_idx == 1 or isinstance(val, str) or pd.isna(val):
                cell.value = "" if pd.isna(val) else str(val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                continue

            # Check if genuinely numeric
            if not isinstance(val, (int, float, np.integer, np.floating)):
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # Numeric Value
            num_val = float(val)
            cell.value = num_val
            cell.alignment = Alignment(horizontal="right", vertical="center")

            # Determine Number Format
            if number_format:
                cell.number_format = number_format
            elif is_metrics_sheet:
                # Percentage metrics: return, cagr, vol, drawdown, var, cvar
                if any(kw in metric_name for kw in ["retorno", "cagr", "volatilidad", "drawdown", "var", "cvar", "rendimiento", "%"]):
                    cell.number_format = "0.00%"
                elif any(kw in metric_name for kw in ["sharpe", "sortino", "calmar", "ratio"]):
                    cell.number_format = "0.000"
                elif "días" in metric_name or "days" in metric_name:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "0.0000"
            elif is_weights_sheet:
                # Weights are percentages
                cell.number_format = "0.00%"
            elif is_wealth_sheet:
                # Dollar currency format for wealth indices
                cell.number_format = "$#,##0.00"
            else:
                if abs(num_val) < 0.001 and num_val != 0.0:
                    cell.number_format = "0.000000"
                elif abs(num_val) < 1.0:
                    cell.number_format = "0.0000"
                else:
                    cell.number_format = "#,##0.00"

    # 3. Total Check Row for Weights Sheet
    if is_weights_sheet and num_rows > 0:
        tot_row = num_rows + 2
        ws.row_dimensions[tot_row].height = 22
        cell_label = ws.cell(row=tot_row, column=1, value="TOTAL")
        cell_label.font = bold_font if bold_font else regular_font
        cell_label.fill = total_fill if total_fill else zebra_fill_gray
        cell_label.alignment = Alignment(horizontal="left", vertical="center")
        cell_label.border = thick_bottom_border if thick_bottom_border else thin_border

        for c_idx in range(2, len(headers) + 1):
            col_letter = get_column_letter(c_idx)
            cell_sum = ws.cell(row=tot_row, column=c_idx)
            cell_sum.value = f"=SUM({col_letter}2:{col_letter}{tot_row - 1})"
            cell_sum.font = bold_font if bold_font else regular_font
            cell_sum.fill = total_fill if total_fill else zebra_fill_gray
            cell_sum.alignment = Alignment(horizontal="right", vertical="center")
            cell_sum.number_format = "0.00%"
            cell_sum.border = thick_bottom_border if thick_bottom_border else thin_border

    # 4. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
